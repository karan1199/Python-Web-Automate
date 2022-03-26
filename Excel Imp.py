import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import RED
from openpyxl.styles.colors import GREEN
from datetime import date
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import ElementClickInterceptedException,ElementNotInteractableException

#Load Todays Sheet
wb = openpyxl.load_workbook('Book1.xlsx')
sheet = wb[str(date.today())]
STATUS = False

def WEB_AUTOMATION(USERNAME,PASSWORD):
    print(USERNAME,PASSWORD)
    ser = Service("chromedriver.exe")
    ser.creationflags = CREATE_NO_WINDOW
    driver=webdriver.Chrome(service=ser)
    #DELAY
    time.sleep(1)
    #FULL SCREEN
    driver.maximize_window()
    #OPEN WEBSITE
    driver.get("https://mastersolutions.online/Home")
    #CLICK DASHBOARD BUTTON
    try:
        driver.find_element(By.XPATH,"/html/body/header/div/div/div[2]/div[2]/a[1]").click()
        time.sleep(1)
        #INPUT USERNAME
        driver.find_element(By.XPATH,"//*[@id=\"UserName\"]").send_keys(USERNAME)
        #PASSWORD
        driver.find_element(By.XPATH,"//*[@id=\"Password\"]").send_keys(PASSWORD)
        #PRESS LOGIN BUTTON
        driver.find_element(By.XPATH,"/html/body/div[2]/div/div/div/div/form/div[3]/input[1]").click()
        time.sleep(2)
        #CHECK FOR BANNER
        try:
            driver.find_element(By.XPATH,"//*[@id=\"NotificationModel\"]/div/div/div[1]/button/span").click()
        except (NoSuchElementException,ElementNotInteractableException,ElementClickInterceptedException):
            pass
        #CLICK TODAYS TASK
        driver.find_element(By.XPATH,"//*[@id=\"sectionsNav\"]/div/div[2]/ul/li[3]/a").click()
        #START THE TASK
        driver.find_element(By.XPATH,"/html/body/div/div[2]/section/section/div[2]/div/div/div/a").click()
    
        #fetch Image and Enter the Code
        while int(driver.find_element(By.XPATH,"/html/body/div/div[2]/section/section/div[2]/div/div[3]/div[2]/div/div[2]/h3").get_attribute('innerHTML')) !=0 :
            src= driver.find_element(By.XPATH,"//*[@id=\"frmJobWork\"]/div[1]/div[1]").get_attribute('innerHTML')
            driver.find_element(By.XPATH,"//*[@id=\"EnteredCaptcha\"]").send_keys(src,Keys.ENTER)
        # IF THE JOB IS DONE
        else:
            if int(driver.find_element(By.XPATH,"/html/body/div/div[2]/section/section/div[2]/div/div[4]/div[1]/div/div[2]/h3").get_attribute('innerHTML')) == int(driver.find_element(By.XPATH,"/html/body/div/div[2]/section/section/div[2]/div/div[3]/div[1]/div/div[2]/h3").get_attribute('innerHTML')):
                return True
            else:
                return False
    except (NoSuchElementException,ElementNotInteractableException,ElementClickInterceptedException):
        return False

for row in range(2,sheet.max_row+1):
    if sheet[str("C" + str(row))].value != 'Done':
        USERNAME = sheet.cell(row=row, column=1).value
        PASSWORD = sheet.cell(row=row, column=2).value
        if USERNAME == None:
            break
        #STATUS = WEB_AUTOMATION(USERNAME,PASSWORD)

        if STATUS == True:
            sheet[str("C" + str(row))]='Done'
            sheet[str("C" + str(row))].fill = PatternFill("solid", fgColor=GREEN)
        elif STATUS == False:
            sheet[str("C" + str(row))]='ERROR'
            sheet[str("C" + str(row))].fill = PatternFill("solid", fgColor=RED)
    STATUS = False

wb.save(filename='Book1.xlsx')


    











#print(sheet.max_row)
#for a1,a2 in sheet[sheet.dimensions]:  
#    print(a1.value, a2.value)
#for row in sheet.iter_rows(min_row=2, min_col=1, max_row=15, max_col=2):
#    Username,Password=row
#    print(Username.value , end="  ")
#    print(Password.value)
   